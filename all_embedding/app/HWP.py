import openpyxl

class HWP():
    def __init__(self, statistic_path, fixed_path, unfix_path, alpha, n, codedoc):
        self.statistic_path = statistic_path
        self.fixed_path = fixed_path
        self.unfixed_path = unfix_path
        self.alpha = alpha
        self.pattern_rank = {}
        self.category_rank = {}
        self.weight = []
        self.rank_list = []
        self.n = n
        self.codedoc = codedoc
    def calc_by_pattern(self):   # read all fixed and unfixed bug (top50?)
        all_pattern_num = {}
        fixed_pattern_num = {}
        weight = {}
        wb = openpyxl.load_workbook(self.statistic_path)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        for i in range(2, ws.max_row+1):
            pattern = ws.cell(i, 2).value
            fixed_pattern_num[pattern] = ws.cell(i, 7).value
            all_pattern_num[pattern] = ws.cell(i, 3).value
            weight[pattern] = (fixed_pattern_num[pattern] * self.alpha) / all_pattern_num[pattern]

        # for data in csv.reader(open(self.fixed_path), 'r'):
        #     if data[1] not in fixed_pattern_num.keys():
        #         fixed_pattern_num[data[1]] = data[2]
        #     else:
        #         fixed_pattern_num[data[1]] += data[2]
        # for pattern in fixed_pattern_num.keys():
        #     weight[pattern] = (fixed_pattern_num[pattern] * self.alpha) / all_pattern_num[pattern]
        self.weight = sorted(weight.items(), key = lambda kv:(kv[1], kv[0]), reverse=True)
    # #
    def calc_by_time(self):
        max = 3600*365*100
        weight = {}
        num = {}
        for doc in self.codedoc:
            pattern = doc.info["violation_type"]
            if pattern not in weight.keys():
                num[pattern] = 1
            else:
                num[pattern] += 1
        for doc in self.codedoc:
            pattern = doc.info["violation_type"]
            if "duration" in doc.info.keys():
                time = doc.info["duration"]
                if pattern not in weight.keys():
                    weight[pattern] = time
                else:
                    weight[pattern] += time
            else:
                if pattern not in weight.keys():
                    weight[pattern] = max
                else:
                    weight[pattern] += max
        for key in weight.keys():
            weight[key] = weight[key]/num[key]
        self.weight = sorted(weight.items(), key=lambda kv: (kv[1], kv[0]))

    def calc_metrics(self, pred_true_bug, pred_false_bug, all_code):
        tp, fp, tn, fn = 0, 0, 0, 0
        for item in pred_true_bug:
            if item.cls == 1:
                tp += 1
            elif item.cls == 0:
                fp += 1
        for item in pred_false_bug:
            if item.cls == 0:
                tn += 1
            elif item.cls == 1:
                fn += 1
        assert tp + fp + tn + fn == len(all_code)
        precision, recall, accuracy = tp/(tp+fp), tp/(tp+fn), (tp+tn)/len(all_code)
        if precision+recall != 0:
            f1 = 2*precision*recall/(precision+recall)
        else:
            f1 = None
        return precision, recall, accuracy, f1

    def prioritize(self, all_code):
        for num in self.n:
            print("top" + str(num) + ":")
            true_bug_pattern = [value[0] for value in self.weight[:num]]
            pred_true_bug = []
            false_bug_pattern = [value[0] for value in self.weight[num:]]
            pred_false_bug = []
            for doc in all_code:
                if doc.info["violation_type"] in true_bug_pattern:
                    pred_true_bug.append(doc)
                elif doc.info["violation_type"] in false_bug_pattern:
                    pred_false_bug.append(doc)
            precision, recall, accuracy, f_score = self.calc_metrics(pred_true_bug, pred_false_bug, all_code)
            print("precision:", precision)
            print("recall:", recall)
            print("accuracy:", accuracy)
            print("f1-score:", f_score)













