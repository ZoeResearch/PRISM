
import pickle
import numpy as np
import os
import time
import random
import collections
import math
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, auc, roc_curve
from sklearn.model_selection import train_test_split
from tensorflow.python.keras.layers import Masking
from tensorflow.python.keras.models import Sequential
from tensorflow.python.keras.layers import LSTM, Dense, GRU, Bidirectional
from tensorflow.python.keras.callbacks import EarlyStopping
from tensorflow.python.keras.layers import Dropout
from tensorflow.python.keras import optimizers
from gensim.models import KeyedVectors
import matplotlib
from tensorflow.python.keras.utils import np_utils
import multiprocessing
# from tensorflow.python.keras.layers import Conv1D
# from tensorflow.python.layers.convolutional import Conv1D
from tensorflow.python.keras import Input
from tensorflow.python.keras.layers import Conv1D, MaxPooling1D, Flatten, BatchNormalization, CuDNNLSTM, CuDNNGRU
matplotlib.use('Agg')
from matplotlib import pyplot as plt
import openpyxl
import argparse
import time
from multiprocessing import Pool
import gc
# import sys
# sys.path.append("../..")
from Util.gen_BYTE_vec import get_byte_embedding, allIns
from Util.gen_IR_vec import get_IR_embedding
import re
import math
from pre_train_def import w2v


CodeDocument = collections.namedtuple('CodeDocument', 'words tags cls info')
CodeDocumentOOP = collections.namedtuple('CodeDocumentOOP', 'words tags cls commitID uuid')

tag = 0
class_dic = {}
class_order = 0


def create_code_document(words, tags, cls, info):
    return CodeDocument(words, [tags], cls, info)


def get_input_shape(conn, args):
    input_shape = None
    if conn == "src" or conn == "ir_seq" or conn == "ir_id_1" or conn == "ir_id_2" or conn == "byte_seq" or conn == "byte_id_1" or conn == "byte_id_2":
        input_shape = (args["sentence_length"], args["voc_size"])
    elif conn == "add" or conn == "sub" or conn == "mul":
        input_shape = (max(args["sentence_length"]), args["voc_size"])
    elif conn == "linear_row":
        input_shape = (args["sentence_length"][0] + args["sentence_length"][1] + args["sentence_length"][2], args["voc_size"])
    elif conn == "linear_col":
        input_shape = (args["sentence_length"], args["voc_size"] * 3)
    elif conn == "cosine_similarity":
        input_shape = (args["sentence_length"], args["sentence_length"])
    elif conn == "all_1":
        input_shape = (args["sentence_length"] * 3, args["voc_size"] * 6)
    elif conn == "all_2":
        input_shape = (args["sentence_length"], args["voc_size"] * 18)
    elif conn == "all_3":
        input_shape = (args["sentence_length"][0],args["voc_size"] * 3 + 3)
    elif conn == "all_4":
        input_shape = (args["sentence_length"],args["voc_size"] * 7)
    elif conn == "3loss_src" or conn == "3loss_ir" or conn == "3loss_byte1" or conn == "3loss_byte2" or conn == "3loss_src_ir" \
            or conn == "3loss_src_byte1" or conn == "3loss_src_byte2" or conn == "3loss_src_byte3" or conn == "3loss_src_ir_byte1" or \
            conn == "3loss_src_ir_byte2":
        input_shape = (250,1)
    elif conn == "pca_src" or conn == "pca_ir" or conn == "pca_byte":
        input_shape = (1,10000)
    return input_shape


def buildlstm_para_rnn(args, flag, classification_model, input_shape):
    model = Sequential()
    if not input_shape:
        print("no input shape!")
        exit(0)

    # model.add(Masking(mask_value=0.0, input_shape=input_shape))
    # model.add(LSTM(args["lstm_unit"], return_sequences=True))
    # model.add(Input(shape=input_shape))
    for i in range(1, args["layer"]):
        if classification_model == "lstm":
            # model.add(LSTM(args["lstm_unit"], return_sequences=True))
            model.add(CuDNNLSTM(args["lstm_unit"], return_sequences=True))
        elif classification_model == "blstm":
            # model.add(Bidirectional(LSTM(args["lstm_unit"], return_sequences=True)))
            model.add(Bidirectional(CuDNNLSTM(args["lstm_unit"], return_sequences=True)))
        elif classification_model == "gru":
            # model.add(GRU(units=args["gru_unit"], activation='tanh', recurrent_activation='sigmoid',return_sequences=True))
            model.add(CuDNNGRU(units=args["gru_unit"], return_sequences=True))
        elif classification_model == "bgru":
            # model.add(Bidirectional(GRU(units=args["gru_unit"], activation='tanh', recurrent_activation='sigmoid', return_sequences=True)))
            model.add(Bidirectional(CuDNNGRU(units=args["gru_unit"], return_sequences=True)))
        model.add(Dropout(args["drop_out"]))
    if classification_model == "lstm":
        # model.add(LSTM(args["lstm_unit"]))
        model.add(CuDNNLSTM(args["lstm_unit"]))
    elif classification_model == "blstm":
        # model.add(Bidirectional(LSTM(args["lstm_unit"])))
        model.add(Bidirectional(CuDNNLSTM(args["lstm_unit"])))
    elif classification_model == "gru":
        # model.add(GRU(units=args["gru_unit"], activation='tanh', recurrent_activation='sigmoid'))
        model.add(CuDNNGRU(units=args["gru_unit"]))
    elif classification_model == "bgru":
        # model.add(Bidirectional(GRU(units=args["gru_unit"], activation='tanh', recurrent_activation='sigmoid')))
        model.add(Bidirectional(CuDNNGRU(units=args["gru_unit"])))
    model.add(Dropout(args["drop_out"]))

    if flag:
        model.add(Dense(args["class_num"], activation='softmax'))
        # model.compile(loss='categorical_crossentropy',
        #                   optimizer=optimizers.adam_v2.Adam(lr=args["learning_rate"]),
        #                   metrics=['accuracy'])
        model.compile(loss='categorical_crossentropy',
                      optimizer=args["optimizer"],
                      metrics=['accuracy'])

    else:
        model.add(Dense(1, activation='sigmoid'))
        # model.compile(loss='binary_crossentropy',
        #               optimizer=optimizers.adam_v2.Adam(lr=args["learning_rate"]),
        #               metrics=["accuracy"])
        model.compile(loss='binary_crossentropy',
                      optimizer=args["optimizer"],
                      metrics=["accuracy"])
    # model.summary()
    return model

def build_para_lr(args, flag, input_shape):
    model = Sequential()
    model.add(Input(shape=input_shape))
    model.add(Flatten())
    if flag:
        model.add(Dense(args["class_num"], activation='softmax'))
        model.compile(loss='categorical_crossentropy',
                      optimizer=optimizers.adam_v2.Adam(lr=args["learning_rate"]),
                      metrics=['accuracy'])

    else:
        model.add(Dense(1, activation='sigmoid'))
        model.compile(loss='binary_crossentropy',
                      optimizer=optimizers.adam_v2.Adam(lr=args["learning_rate"]),
                      metrics=["accuracy"])

    model.summary()
    return model

def build_para_mlp(args, flag, input_shape):
    model = Sequential()
    model.add(Input(shape=input_shape))
    for i in range(args["layer"]):
        model.add(Dense(args["dense_unit"], activation='relu'))
    model.add(Flatten())
    if flag:
        model.add(Dense(args["class_num"], activation='softmax'))
        model.compile(loss='categorical_crossentropy',
                      optimizer=optimizers.adam_v2.Adam(lr=args["learning_rate"]),
                      metrics=['accuracy'])

    else:
        model.add(Dense(1, activation='sigmoid'))
        model.compile(loss='binary_crossentropy',
                      optimizer=optimizers.adam_v2.Adam(lr=args["learning_rate"]),
                      metrics=["accuracy"])

    model.summary()
    return model

def add_output_layer(flag, args, model):
    if flag:
        model.add(Dense(args["class_num"], activation='softmax'))
        model.compile(loss='categorical_crossentropy',
                      optimizer=optimizers.adam_v2.Adam(lr=args["learning_rate"]),
                      metrics=['accuracy'])

    else:
        model.add(Dense(1, activation='sigmoid'))
        model.compile(loss='binary_crossentropy',
                      optimizer=optimizers.adam_v2.Adam(lr=args["learning_rate"]),
                      metrics=["accuracy"])
    return model

def build_para_cnn(args, flag, input_shape):
    model = Sequential()
    model.add(Conv1D(256, args["kernel_size"], padding='same', input_shape = input_shape))
    model.add(MaxPooling1D(args["pool_size"], args["pool_size"], padding='same'))
    model.add(Conv1D(128, args["kernel_size"], padding='same'))
    model.add(MaxPooling1D(args["pool_size"], args["pool_size"], padding='same'))
    model.add(Conv1D(64, args["kernel_size"], padding='same'))
    model.add(Flatten())
    model.add(Dropout(args["drop_out"]))
    model.add(BatchNormalization())  # (批)规范化层
    model.add(Dense(256, activation='relu'))
    model.add(Dropout(args["drop_out"]))
    model = add_output_layer(flag, args, model)
    model.summary()
    return model


def read_doc(doc_path):
    f = open(doc_path, "rb")
    all_code_blocks = []
    all_labels = []
    for block in pickle.load(f):
        # if block.words:
        #     try:
        #         " ".join(block.words).encode(encoding='UTF-8', errors='strict')
        #         code = list(filter(None, [i.replace("\n", "") for i in block.words]))
        #         all_code_blocks.append(code)
        #         all_labels.append(block.cls)
        #     except UnicodeEncodeError:
        #         print(block.words)
        all_code_blocks.append(block.words)
        all_labels.append(block.cls)
    return all_code_blocks, all_labels

def getIRIDF(k, all_code):
    all_doc_num = 0
    opcodeToBlockCounts = {}
    opcodeIDF = {}
    for i in range(k):
        all_doc_num += len(all_code[i])
        for doc in all_code[i]:
            opcode_set = set()
            for line in doc.words:
                for keys, values in line["operator"].items():
                    # opcode = ""
                    for value in values:
                        opcode_set.add(value)
            for opcode in opcode_set:
                if opcode not in opcodeToBlockCounts.keys():
                    opcodeToBlockCounts[opcode] = 1
                else:
                    opcodeToBlockCounts[opcode] += 1
    for key in opcodeToBlockCounts.keys():
        opcodeIDF[key] = np.log10(all_doc_num/opcodeToBlockCounts[key])
    return opcodeIDF

def getByteIDF(k, all_code):
    all_doc_num = 0
    opcodeToBlockCounts = {}
    opcodeIDF = {}
    for i in range(k):
        all_doc_num += len(all_code[i])
        for doc in all_code[i]:
            opcode_set = set()
            for token in doc.words:
                if token in allIns:
                    opcode_set.add(token)
            for opcode in opcode_set:
                if opcode not in opcodeToBlockCounts.keys():
                    opcodeToBlockCounts[opcode] = 1
                else:
                    opcodeToBlockCounts[opcode] += 1
    for key in opcodeToBlockCounts.keys():
        opcodeIDF[key] = np.log10(all_doc_num/opcodeToBlockCounts[key])

    return opcodeIDF

def process_all_data(pre_model_path, code, embed_arg, cls, doc_path, rank_file, K_fold, mul_bin_flag):
    # vec_path = cls + "/" + "vec_" + str(embed_arg['iter']) + "_" + str(
    #     embed_arg['window']) + "_" + str(embed_arg['voc_size'])
    # label_path = cls + "/" + "label_" + str(embed_arg['iter']) + "_" + str(
    #     embed_arg['window']) + "_" + str(embed_arg['voc_size'])
    print("Loading model...")
    model = KeyedVectors.load(pre_model_path, mmap="r")
    all_vec_part, all_label_part = prepare_data(doc_path, model, embed_arg["voc_size"],
                                                embed_arg["sentence_length"], code, rank_file, K_fold,
                                                mul_bin_flag)
    print("Data load finished!")
    return all_vec_part, all_label_part

def split_codedoc(doc_path, rank_base, k, mul_bin_flag):
    all_code = []
    for i in range(k):
        all_code.append([])
    if mul_bin_flag == 0:
        file_name_list = ["good_new", "bad_new"]
        # file_name_list = ["good", "bad"]
    elif mul_bin_flag == 1:
        file_name_list = ["norm", "MT_CORRECTNESS", "BAD_PRACTICE", "CORRECTNESS", "PERFORMANCE", "STYLE", "SECURITY", "I18N", "EXPERIMENTAL"]
        # file_name_list = ["MT_CORRECTNESS", "BAD_PRACTICE", "CORRECTNESS", "PERFORMANCE", "STYLE", "SECURITY", "I18N", "EXPERIMENTAL"]
    for file in file_name_list:
        shuffle_contents = []
        rank_file_path = os.path.join(rank_base, file)
        file_path = os.path.join(doc_path, file)
        print(file_path)
        contents = pickle.load(open(file_path, "rb"))
        # print(contents[0])
        print(file + " length:", len(contents))
        if os.path.exists(rank_file_path):
            rank_file = open(rank_file_path, "rb")
            rank = pickle.load(rank_file)
        else:
            rank_file = open(rank_file_path, "wb")
            rank = [i for i in range(len(contents))]
            random.shuffle(rank)
            pickle.dump(rank, rank_file)
        print("rank length:", len(rank))
        for i in rank:
            shuffle_contents.append(contents[i])
        part_length = len(shuffle_contents) // k
        for i in range(k):
            j = i * part_length
            while j >= i*part_length and j < (i+1)*part_length:
                all_code[i].append(shuffle_contents[j])
                j += 1
    return all_code

def prepare_data(doc_path, model, voc_size, sentence_length, code, rank_base, k, mul_bin_flag):
    length = []
    all_vec_part = []
    all_label_part = []

    ignore_list = ['if_icmpeq', 'if_icmpne', 'if_icmplt', 'if_icmpge', 'if_icmpgt', 'if_icmple', 'if_acmpeq',
                   'if_acmpne', 'goto_w', 'jsr_w']

    all_code = split_codedoc(doc_path, rank_base, k, mul_bin_flag)

    if code == "src":
        for i in range(k):
            all_vec_part.append(get_vec_concat(model, all_code[i], voc_size, sentence_length, allIns, ignore_list, regulate_byte_flag="False"))
            length.extend(count_length(all_vec_part[i]))
            all_label_part.append(get_label(all_code[i], mul_bin_flag))
    elif code == "byte_seq":
        for i in range(k):
            all_vec_part.append(get_vec_concat(model, all_code[i], voc_size, sentence_length, allIns, ignore_list, regulate_byte_flag="True"))
            # all_vec_part[i] = get_vec_concat(model, all_code[i], voc_size, sentence_length)
            length.extend(count_length(all_vec_part[i]))
            all_label_part.append(get_label(all_code[i], mul_bin_flag))
    elif code == "byte_id_1":
        codeIDF = getByteIDF(k, all_code)
        for i in range(k):
            all_code_blocks = get_code(all_code[i])
            all_vec_part.append(get_byte_embedding(model, all_code_blocks, voc_size, codeIDF, code))
            length.extend(count_length(all_vec_part[i]))
            all_vec_part[i] = get_vec_fix(all_vec_part[i], sentence_length, voc_size)
            all_label_part.append(get_label(all_code[i], mul_bin_flag))
    elif code == "byte_id_2":
        codeIDF = getByteIDF(k, all_code)
        for i in range(k):
            all_code_blocks = get_regulate_code(all_code[i], allIns, ignore_list)
            all_vec_part.append(get_byte_embedding(model, all_code_blocks, voc_size, codeIDF, code))
            length.extend(count_length(all_vec_part[i]))
            all_vec_part[i] = get_vec_fix(all_vec_part[i], sentence_length, voc_size)
            all_label_part.append(get_label(all_code[i], mul_bin_flag))
    elif code == "byte_id_3":
        codeIDF = getByteIDF(k, all_code)
        for i in range(k):
            all_code_blocks = get_regulate_code(all_code[i], allIns, ignore_list)
            all_vec_part.append(get_byte_embedding(model, all_code_blocks, voc_size, codeIDF, code))
            length.extend(count_length(all_vec_part[i]))
            all_vec_part[i] = get_vec_fix(all_vec_part[i], sentence_length, voc_size)
            all_label_part.append(get_label(all_code[i], mul_bin_flag))
    elif code == "ir_seq":
        for i in range(k):
            all_vec_part.append(get_ir_vec_fix(model, all_code[i], voc_size, sentence_length))
            length.extend(count_length(all_vec_part[i]))
            all_label_part.append(get_label(all_code[i], mul_bin_flag))
    elif code == "ir_id_1" or code == "ir_id_2":
        codeIDF = getIRIDF(k, all_code)
        for i in range(k):
            all_vec_part.append(get_IR_embedding(model, all_code[i], voc_size, sentence_length, code, codeIDF))
            length.extend(count_length(all_vec_part[i]))
            all_vec_part[i] = get_vec_fix(all_vec_part[i], sentence_length, voc_size)
            all_label_part.append(get_label(all_code[i], mul_bin_flag))
    print("average length:", sum(length) / len(length))
    print("length of every part:")
    for i in range(len(all_vec_part)):
        print(len(all_vec_part[i]))
    return all_vec_part, all_label_part

def split_dataset(all_code_part, all_label_part, class_num, times, k, split_flag, categorical_flag):
    x_train, y_train = [], []
    if split_flag == "True":
        for i in range(1, k):
        # for i in range(0, k-1):
            if not isinstance(all_code_part[i], list):
                all_code_part[i] = all_code_part[i].tolist()
            if i-1 == times:
            # if i == times:
                x_val = np.asarray(all_code_part[i]).astype(np.float32)
                y_val = np.asarray(all_label_part[i])
                # x_val = np.asarray(all_code_part[i]).astype(np.float16)
                # y_val = np.asarray(all_label_part[i])
            else:
                x_train += all_code_part[i]
                y_train += all_label_part[i]

        x_test, y_test = np.asarray(all_code_part[0]).astype(np.float32), np.asarray(all_label_part[0])
        # x_test, y_test = np.asarray(all_code_part[0]).astype(np.float16), np.asarray(all_label_part[0])
        # x_test, y_test = np.asarray(all_code_part[k-1]).astype(np.float32), np.asarray(all_label_part[k-1]).astype(np.float32)

    elif split_flag == "False":
        for i in range(k):
            if not isinstance(all_code_part[i], list):
                all_code_part[i] = all_code_part[i].tolist()
            if i == times:
                x_val = np.asarray(all_code_part[i]).astype(np.float32)
                y_val = np.asarray(all_label_part[i])
                # x_val = np.asarray(all_code_part[i]).astype(np.float16)
                # y_val = np.asarray(all_label_part[i]).astype(np.float16)
            else:
                x_train += all_code_part[i]
                y_train += all_label_part[i]
        x_test, y_test = x_val, y_val

    x_train = np.asarray(x_train).astype(np.float32)
    # x_train = np.asarray(x_train).astype(np.float16)
    if class_num == 2 or categorical_flag == "False":
        y_train = np.asarray(y_train)
        # y_train = np.asarray(y_train).astype(np.float16)
    elif categorical_flag == "True":
        y_train = np_utils.to_categorical(np.asarray(y_train), num_classes=class_num)
        # y_train = np_utils.to_categorical(np.asarray(y_train).astype(np.float16), num_classes=class_num)
    return x_train, y_train, x_val, y_val, x_test, y_test

def single_get_vec_concat(doc, voc_size, sentence_length, model):
    vec = []
    if doc.words:
        code = list(filter(None, doc.words))
        for word in code:
            try:
                if len(vec) >= sentence_length:
                    break
                vec.append(model[word])
            except KeyError as e:
                print(e)
                continue
        while len(vec) < sentence_length:
            vec.append(np.zeros(voc_size))
    return vec

def get_vec_concat(model, all_codeDoc, voc_size, sentence_length, operator_set, ignore_list, regulate_byte_flag):
    # print(model[","])
    all_vec = []
    num = 0
    num1 = 0
    for doc in all_codeDoc:
        vec = []
        if doc.words:
            if regulate_byte_flag == "True":
                code = regulate(doc.words, operator_set, ignore_list)
            else:
                code = list(filter(None, doc.words))
            for word in code:
                try:
                    if len(vec) >= sentence_length:
                        break
                    # vec.append(model[word].astype(np.float16))
                    vec.append(model[word])
                except KeyError as e:
                    print(e)
                    # print(doc)
                    continue
            while len(vec) < sentence_length:
                # vec.append(np.zeros(voc_size).astype(np.float16))
                vec.append(np.zeros(voc_size))
            all_vec.append(vec)
        else:
            all_vec.append(vec)
            num1 += 1
    # print("oov number:", num)
    print("null block", num1)
    return all_vec

def get_vec_fix(vectors, sentence_length, voc_size):
    all_vec = []
    for sentence_vec in vectors:
        vec = []
        for word_vec in sentence_vec:
            if len(vec) >= sentence_length:
                break
            vec.append(word_vec)
        while len(vec) < sentence_length:
            vec.append(np.zeros(voc_size))
        all_vec.append(vec)
    return all_vec

def count_length(vectors):
    length = []
    for sentence_vec in vectors:
        length.append(len(sentence_vec))
    return length

def get_ir_vec_fix(model, all_codeDoc, voc_size, sentence_length):
    all_vec = []
    for block in all_codeDoc:
        if block.words:
            code = []
            vec = []
            j = 0
            while j+1<len(block.words):
                if block.words[j] == ":" and block.words[j+1] == "=":
                    code.append(":=")
                    j += 2
                else:
                    code.append(block.words[j])
                    j += 1
            for word in code:
                try:
                    if len(vec) >= sentence_length:
                        break
                    vec.append(list(model[word]))
                except KeyError as e:
                    print(e)
                    print(block)
                    continue
            while len(vec) < sentence_length:
                vec.append(list(np.zeros(voc_size)))
            all_vec.append(vec)

    return all_vec

def get_label(all_codeDoc, mul_bin_flag):
    all_label = []
    if mul_bin_flag == 0:
        for doc in all_codeDoc:
            if doc.words:
                if doc.cls != 0:
                    all_label.append(1)
                else:
                    all_label.append(0)
    elif mul_bin_flag == 1:
        for doc in all_codeDoc:
            if doc.words:
                all_label.append(doc.cls)
    return all_label


def get_code(all_codeDoc):
    all_code = []
    for doc in all_codeDoc:
        if doc.words:
            all_code.append(doc.words)
    return all_code

def regulate(words, operator_set, ignore_list):
    code = []
    for token in words:
        if token in operator_set and "_" in token and token not in ignore_list:
            code.append(token.split("_")[0])
            code.append(token.split("_")[1])
        else:
            code.append(token)
    return code

def get_regulate_code(blocks, operator_set, ignore_list):
    all_code_blocks = []
    for block in blocks:
        if block.words:
            code = regulate(block.words, operator_set, ignore_list)
            all_code_blocks.append(code)
    return all_code_blocks


def count_cls(doc_IR):
    num = 0
    max_num = -1
    cls_before = doc_IR[0].cls
    for i in doc_IR:
        if cls_before == i.cls:
            num = num + 1
            cls_before = i.cls
        else:
            max_num = max(max_num, num)
            num = 0
            cls_before = i.cls
    return max_num


def get_score_multiclassfication(y_pred, y_true, class_num):
    n_class = class_num
    accuracy = accuracy_score(y_true, y_pred)
    micro_precision = precision_score(y_true, y_pred, average='micro')
    macro_precision = precision_score(y_true, y_pred, average='macro')
    weight_precision = precision_score(y_true, y_pred, average='weighted')
    micro_recall = recall_score(y_true, y_pred, average='micro')
    macro_recall = recall_score(y_true, y_pred, average='macro')
    weight_recall = recall_score(y_true, y_pred, average='weighted')
    micro_f1 = f1_score(y_true, y_pred, average='micro')
    macro_f1 = f1_score(y_true, y_pred, average='macro')
    weight_f1 = f1_score(y_true, y_pred, average='weighted')
    fpr = dict()
    tpr = dict()
    roc_auc = dict()
    y_pred_roc = np_utils.to_categorical(y_pred, num_classes=class_num)
    y_true_roc = np_utils.to_categorical(y_true, num_classes=class_num)
    for i in range(n_class):
        fpr[i], tpr[i], _ = roc_curve(y_true_roc[:, i], y_pred_roc[:, i])
        roc_auc[i] = auc(fpr[i], tpr[i])
    fpr["micro"], tpr["micro"], _ = roc_curve(y_true_roc.ravel(), y_pred_roc.ravel())
    roc_auc["micro"] = auc(fpr["micro"], tpr["micro"])

    # Compute macro-average ROC curve and ROC area（方法一）
    # First aggregate all false positive rates
    all_fpr = np.unique(np.concatenate([fpr[i] for i in range(class_num)]))
    # Then interpolate all ROC curves at this points
    mean_tpr = np.zeros_like(all_fpr)
    for i in range(n_class):
        mean_tpr += np.interp(all_fpr, fpr[i], tpr[i])
    # Finally average it and compute AUC
    mean_tpr /= n_class
    fpr["macro"] = all_fpr
    tpr["macro"] = mean_tpr
    roc_auc["macro"] = auc(fpr["macro"], tpr["macro"])
    score = dict(accuracy=accuracy, micro_precision=micro_precision, macro_precision=macro_precision, weight_precision=weight_precision, micro_recall=micro_recall, macro_recall=macro_recall, weight_recall=weight_recall,
                 micro_f1=micro_f1, macro_f1=macro_f1, weight_f1=weight_f1, micro_fpr=fpr["micro"], macro_fpr=fpr["macro"], micro_tpr=tpr["micro"], macro_tpr=tpr["macro"], micro_auc=roc_auc["micro"], macro_auc=roc_auc["macro"])
    return score


def get_score_binaryclassfication(y_pred, y_true):
    accuracy = accuracy_score(y_true, y_pred)
    precision = precision_score(y_true, y_pred)
    recall = recall_score(y_true, y_pred, average='binary')
    f1 = f1_score(y_true, y_pred)
    fpr, tpr, thresholds = roc_curve(y_pred, y_true)
    auc_score = auc(fpr, tpr)
    score = dict(accuracy=accuracy, precision=precision, recall=recall, f1=f1, fpr=fpr, tpr=tpr, thresholds=thresholds, auc_score=auc_score)
    return score

def prob_mean(prob2label, prob2index, times):
    bug = []
    norm = []
    for i in sorted(prob2label):
        if prob2label[i] == 1:
            bug.append(i)
        else:
            norm.append(i)
    # plt.figure()
    # # x1 = np.linspace(0, len(bug)-1, len(bug))
    # x1 = np.linspace(0, 7000, len(bug))
    # y1 = bug
    # plt.plot(x1, y1)
    # # plt.grid(True)
    # # plt.savefig(str(times)+"_bug_prob.jpg"
    # # plt.figure()
    # # x2 = np.linspace(0, len(norm)-1, len(norm))
    # x2 = np.linspace(0, 7000, len(norm))
    # y2 = norm
    # plt.plot(x2, y2)
    # plt.grid(True)
    # plt.savefig(str(times)+"prob.jpg")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "bug"
    ws.cell(1, 2).value = "norm"
    for i in range(len(bug)):
        ws.cell(i + 2, 1).value = bug[i]
    for j in range(len(norm)):
        ws.cell(j + 2, 2).value = norm[j]
    wb.save(str(times)+"prob.xlsx")

    wb2 = openpyxl.Workbook()
    ws = wb2.active
    ws.cell(1, 1).value = "prob"
    ws.cell(1, 2).value = "index"
    prob = sorted(prob2index)[::-1]
    for i in range(len(prob)):
        ws.cell(i + 2, 1).value = prob[i]
        ws.cell(i + 2, 2).value = prob2index[prob[i]]

    wb2.save(str(times) + "index.xlsx")

    return sum(bug)/len(bug), sum(norm)/len(norm)

def add_metric(detect_model, test_X, test_y, test_score, times):
    y_prob = detect_model.predict(test_X).tolist()
    prob2label = {}
    prob2index = {}
    for i in range(len(y_prob)):
        prob2index[y_prob[i][0]] = i
        prob2label[y_prob[i][0]] = test_y[i]
    test_score["bug_prob"], test_score["norm_prob"] = prob_mean(prob2label, prob2index, times)
    for k in [10, 50, 100, 150, 200]:
        print("*************** top " + str(k) + " metrics: ****************")
        TP, top_k_precision, top_k_recall = get_top_k_metrics(k, prob2label)
        print("top_k_TP:", str(TP))
        print("top_k_precision:", str(top_k_precision))
        print("top_k_recall:", str(top_k_recall))
        test_score["top_" + str(k) + "_TP"] = TP
        test_score["top_" + str(k) + "_precision"] = top_k_precision
        test_score["top_" + str(k) + "_recall"] = top_k_recall
    return test_score

def get_top_k_metrics(k, prob2label):
    TP = 0
    FP = 0
    FN = 0
    for i in sorted(prob2label)[::-1][:math.floor(k)]:
        if i > 0.5 and prob2label[i] == 1:
            TP += 1
        elif i > 0.5 and prob2label[i] == 0:
            FP += 1
        elif i <= 0.5 and prob2label[i] == 1:
            FN += 1
    if TP+FP != 0:
        top_k_precision = TP/(TP+FP)
    else:
        top_k_precision = None
    if TP+FN != 0:
        top_k_recall = TP/(TP+FN)
    else:
        top_k_recall = None
    return TP, top_k_precision, top_k_recall

def makedata_s(ori_vec, y, class_num):
    # Data generated here is used for CNN model not textcnn
    X = ori_vec.reshape(ori_vec.shape[0], 1, ori_vec.shape[1])

    all_data_len = len(X)
    rank_filename = "../../pickle_object/rank" + str(all_data_len)
    if os.path.exists(rank_filename):
        rank_file = open(rank_filename, "rb")
        rank = pickle.load(rank_file)
    else:
        rank_file = open(rank_filename, "wb")
        rank = [i for i in range(all_data_len)]
        random.shuffle(rank)
        pickle.dump(rank, rank_file)

    X_after_shuffle = []
    y_after_shuffle = []
    for i in rank:
        X_after_shuffle.append(X[i])
        y_after_shuffle.append(y[i])
    num_train = int(all_data_len // 1.25)
    X_after_shuffle = np.array(X_after_shuffle)
    y_after_shuffle = np.array(y_after_shuffle)
    x_train, x_test = X_after_shuffle[:num_train], X_after_shuffle[num_train:]
    y_train, y_test = y_after_shuffle[:num_train], y_after_shuffle[num_train:]
    x_train, x_val, y_train, y_val = train_test_split(x_train, y_train, test_size=0.1, random_state=42)
    y_train = np_utils.to_categorical(y_train, num_classes=class_num)
    return x_train, x_val, x_test, y_train, y_val, y_test


def Merge(dict1, dict2):
    res = {**dict1, **dict2}
    return res


def write_record(args, score, file_name):
    if os.path.exists(file_name):
        f = open(file_name, "a")
    else:
        f = open(file_name, "a")
        for i in args.keys():
            f.write(str(i))
            f.write("\t")
        for i in score.keys():
            f.write(str(i))
            f.write("\t")
        f.write("\n")
    for i in args.values():
        f.write(str(i))
        f.write("\t")
    for i in score.values():
        f.write(str(i))
        f.write("\t")
    f.write("\n")
    f.close()


def write_best_score(best_args, test_score, conn):
    f = open("best_record_" + conn, "w")
    for i in best_args.keys():
        f.write(str(i))
        f.write("\t")
    f.write("\n")
    for i in best_args.values():
        f.write(str(i))
        f.write(" ")
    f.write("\n\n")
    f.write("Get score:\n")

    for i in test_score.keys():
        f.write(str(i))
        f.write("\t")
    f.write("\n")
    for i in test_score.values():
        f.write(str(i))
        f.write("\t")
    f.write("\n")
    f.close()


def check_exist():
    if os.path.exists("./score_record"):
        os.remove("./score_record")
    if os.path.exists("./best_record"):
        os.remove("./best_record")

def dump_object(path, vec_doc):
    if os.path.exists(path):
        return
    f = open(path, "wb")
    pickle.dump(vec_doc, f)
    f.close()

def plot_figure(history, path, times):
    # Plot history: MAE
    # plt.figure()
    # plt.plot(history.history['loss'], label='MAE (training data)')
    # plt.plot(history.history['val_loss'], label='MAE (validation data)')
    # plt.title('loss')
    # plt.ylabel('MAE value')
    # plt.xlabel('No. epoch')
    # plt.legend(loc="upper left")
    # # plt.savefig('..\\fig\\loss.jpg')
    # plt.savefig(path + "loss" + str(times) +".jpg")
    # # plt.show()

    plt.figure()
    plt.plot(history.history['acc'], label='Acc (training data)')
    plt.plot(history.history['val_acc'], label='Acc (validation data)')
    plt.title('accuracy')
    plt.ylabel('Accuracy value')
    plt.xlabel('No. epoch')
    plt.legend(loc="upper left")
    # plt.savefig('..\\fig\\accuracy.jpg')
    plt.savefig(path + "acc" + str(times) +".jpg")
    # plt.show()